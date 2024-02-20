# -*- coding: utf-8 -*-
"""
@author: JOHAN RUJANO
"""
import argparse  # noqa: F401
import configparser  # noqa: F401
import csv  # noqa: F401
import os  # noqa: F401

from openpyxl import (  # noqa: F401
    Workbook,  # noqa: F401
    worksheet,
)
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: F401

from Extractor.extractor_interface import SnowflakeExtractor  # noqa: F401


def render_total_employee(
    ws1: worksheet,
    label_total: str,
    sum_vta_employee: float,
    sum_pro_employee: float,
    sum_tot_employee: float,
    r: int
) -> worksheet:
    header_font = Font(size=12, bold=True)
    columnas_total = range(6, 10)
    dtotal = [
        "",
        "",
        "",
        "",
        "",
        f"Total {label_total}",
        sum_vta_employee,
        sum_pro_employee,
        sum_tot_employee,
    ]
    ws1.append(dtotal)
    for columna in columnas_total:
        cell = ws1.cell(row=r, column=columna)
        cell.font = header_font
        cell.number_format = '#,##0.' 
    return ws1


def render_title_columns(
    ws1: worksheet,
    col_header: list,
    r: int
) -> worksheet:
    header_font = Font(size=12, bold=True)
    columnas_titulos = range(1, 10)
    ws1.append(col_header)
    for columna in columnas_titulos:
        cell = ws1.cell(row=r, column=columna)
        cell.font = header_font
    return ws1

def format_number(
    ws1 : worksheet) -> worksheet:
    columnas_total = ['G', 'H', 'I']
    for columna in columnas_total:
        columna_number = ws1[columna]
        for celda in columna_number:
            celda.number_format = '#,##0.'
    return ws1

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generador de reportes Ventas Con propinas TBK.")
    parser.add_argument("-r", "--rut", type=str, help="RUT", required=True)
    parser.add_argument(
        "-f", "--date_process", type=str, help="Fecha proceso", required=True
    )
    args = parser.parse_args()
    print("Ventas con propinas")

    query = """SELECT
                tar.RUT_CLIENTE, tar.ID_TRANSACCION, tx_ptlf.cdg_comercio, tx_ptlf.id_terminal, 
                to_varchar( to_date(tx_ptlf.tms_transaccion), 'DD-MM-YYYY' ) AS "Fecha",
                to_time(tx_ptlf.tms_transaccion)::varchar as "Hora",
                tx_ptlf.cdg_autorizacion as "Autorizacion",
                tx_ptlf.cdg_empleado_dependiente as "Empleado",
                (tx_ptlf.mto_transaccion - tx_ptlf.mto_propina) as "Monto Venta",
                (tx_ptlf.mto_propina) as "Monto Propina",
                (tx_ptlf.mto_transaccion) as "Monto Total"
            FROM
                CONSUMER_PROD.BUT.VW_TX_TARIFICADA tar 
                INNER JOIN
                CONSUMER_PROD.BUT.VW_TX_PTLF tx_ptlf ON tar.id_transaccion = tx_ptlf.id_transaccion
            -- WHERE  to_date(tar.tms_transaccion) = to_date ('2024-02-18')
            WHERE   tx_ptlf.MTO_PROPINA > 0
                    and tx_ptlf.CDG_MONEDA in (152, 850)
                    and tx_ptlf.CDG_TIPO_TX <> 104
                    and tar.RUT_CLIENTE='763788318'
            ORDER BY  cdg_empleado_dependiente ASC
            LIMIT 100
            """
    if args.rut:
        query = query.replace('763788318', str(args.rut))
    if args.date_process:
        query = query.replace('2024-02-18', str(args.date_process))

    se = SnowflakeExtractor()
    df = se.extract_simple_df(request_query=query)
    df_columns = df.columns.values.tolist()
    
    # csv
    # print(df_columns)
    path_file_csv = os.path.join(
            "temps", f"propinas_{args.rut}.csv"
    )
    print(path_file_csv)
    gb = df.groupby(str('Empleado'))
    # for name_of_group, contents_of_group in gb:
    #     # print(name_of_group)
    #     print(contents_of_group)

    # for k, gr in gb:
    #     # do your stuff instead of print
    #     print(k)
    #     print(type(gr)) # This will output <class 'pandas.core.frame.DataFrame'>
    #     print(gr)
    #     # You can save each 'gr' in a csv as follows
    #     gr.to_csv('{}.csv'.format(k))

    df_columns = [
        "Código Comercio",
        "Equipo DDLL",
        "Fecha de Transacción",
        "Hora de Transacción",
        "Código de autorización",
        "Código de empleado",
        "Monto Venta",
        "Monto Propina",
        "Monto Total",
    ]
    with open(path_file_csv, 'w', newline='') as file:
        writer = csv.writer(file, delimiter=';', quotechar=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(df_columns)
        # writer.writerow(['name', 'age'])
        # writer.writerow(['John Doe', 30])
        cdg_employee = None
        sum_vta = 0
        sum_pro = 0
        sum_tot = 0
        for i in range(len(df)):
            # print(df.loc[i, "RUT_CLIENTE"], df.loc[i, "ID_TRANSACCION"])
            if cdg_employee is None:
                sum_vta_employee = 0
                sum_pro_employee = 0
                sum_tot_employee = 0
                cdg_employee = df.loc[i, "Empleado"].strip()

            elif cdg_employee != df.loc[i, "Empleado"]:
                writer.writerow(["", "", "", "", "", f"Total Empleado {cdg_employee}", sum_vta_employee, sum_pro_employee, sum_tot_employee ])
                writer.writerow([])
                writer.writerow([])
                writer.writerow(df_columns)
                cdg_employee = df.loc[i, "Empleado"]
                sum_vta_employee = 0
                sum_pro_employee = 0
                sum_tot_employee = 0

            sum_vta += df.loc[i, "Monto Venta"]
            sum_pro += df.loc[i, "Monto Propina"]
            sum_tot += df.loc[i, "Monto Total"]
            sum_vta_employee += df.loc[i, "Monto Venta"]
            sum_pro_employee += df.loc[i, "Monto Propina"]
            sum_tot_employee += df.loc[i, "Monto Total"]

            # print(f"Empleado {df.loc[i, "Empleado"]}")

            writer.writerow([df.loc[i, "CDG_COMERCIO"], df.loc[i, "ID_TERMINAL"], df.loc[i, "Fecha"], df.loc[i, "Hora"], df.loc[i, "Autorizacion"], df.loc[i, "Empleado"], df.loc[i, "Monto Venta"], df.loc[i, "Monto Propina"], df.loc[i, "Monto Total"]])
            if i == len(df)-1:
                writer.writerow(["", "", "", "", "", f"Total Empleado {cdg_employee}", sum_vta_employee, sum_pro_employee, sum_tot_employee ])
                writer.writerow([])
        writer.writerow(["", "", "", "", "", f"Total {args.rut}", sum_vta, sum_pro, sum_tot])
        # for name_of_group, contents_of_group in gb:
        #     # print(name_of_group)
        #     writer.writerow([(contents_of_group)])


    # Excel
    wb = Workbook()
    # work with default worksheet
    ws1 = wb.active  
    # add name
    ws1.title = args.rut
    # ws1 = wb.create_sheet(args.rut, 0)
    col_header = [r for r in df_columns]  # List of column headers
    ws1 = render_title_columns(ws1, col_header, 1)
    r, c = 2, 0  # row=2 and column=0
    cdg_employee = None
    sum_vta = 0
    sum_pro = 0
    sum_tot = 0
    for i in range(len(df)):
        if cdg_employee is None:
            sum_vta_employee = 0
            sum_pro_employee = 0
            sum_tot_employee = 0
            cdg_employee = df.loc[i, "Empleado"]
        elif cdg_employee != df.loc[i, "Empleado"]:
            ws1 = render_total_employee(ws1, f"Empleado: {cdg_employee}", sum_vta_employee, sum_pro_employee, sum_tot_employee, r)
            ws1.append([])
            ws1.append([])
            r += 3
            ws1 = render_title_columns(ws1, col_header, r)
            r += 1
            cdg_employee = df.loc[i, "Empleado"]
            sum_vta_employee = 0
            sum_pro_employee = 0
            sum_tot_employee = 0

        sum_vta += df.loc[i, "Monto Venta"]
        sum_pro += df.loc[i, "Monto Propina"]
        sum_tot += df.loc[i, "Monto Total"]
        sum_vta_employee += df.loc[i, "Monto Venta"]
        sum_pro_employee += df.loc[i, "Monto Propina"]
        sum_tot_employee += df.loc[i, "Monto Total"]

        d = [
            df.loc[i, "CDG_COMERCIO"],
            df.loc[i, "ID_TERMINAL"],
            df.loc[i, "Fecha"],
            df.loc[i, "Hora"],
            df.loc[i, "Autorizacion"],
            df.loc[i, "Empleado"],
            float(df.loc[i, "Monto Venta"]),
            float(df.loc[i, "Monto Propina"]),
            float(df.loc[i, "Monto Total"]),
        ]

        ws1.append(d)
        r += 1
        # Termina el grupo y el ultimo registro
        if i == len(df) - 1:
            ws1 = render_total_employee(ws1, f"Empleado: {cdg_employee}", sum_vta_employee, sum_pro_employee, sum_tot_employee, r)
            ws1.append([])
            ws1.append([])
            r += 3
    ws1 = format_number(ws1)
    ws1 = render_total_employee(ws1, f"R.U.T.: {args.rut}", sum_vta, sum_pro, sum_tot, r)
    path_file_xlsx = os.path.join("temps", f"propinas_{args.rut}.xlsx")
    wb.save(path_file_xlsx)
